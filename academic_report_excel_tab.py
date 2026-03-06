"""
academic_report_excel_tab.py
============================
Enhanced Academic Report Tab for Teacher Pehpeh — IBT
Reads student data directly from the IBT Grade Tracker Excel (Roster + subject sheets).
Falls back to session_state grade_history when no Excel is uploaded.

Integration into app.py:
  At top of file, add:
      try:
          from academic_report_excel_tab import render_academic_report_from_excel
          EXCEL_REPORT_AVAILABLE = True
      except ImportError:
          EXCEL_REPORT_AVAILABLE = False

  In Tab 5 block, replace contents with:
      if EXCEL_REPORT_AVAILABLE:
          render_academic_report_from_excel()
      else:
          # (keep original grade_history-based code as fallback)

Built by IBT | Teacher Pehpeh Project
"""

import streamlit as st
import io
import datetime

# ── IBT Brand Colors ────────────────────────────────────────────────────────
C_NAVY  = "#0F2247"
C_GOLD  = "#D4A843"
C_RED   = "#EF5350"
C_GREEN = "#66BB6A"
C_BLUE  = "#2B7DE9"
C_AMBER = "#FFA726"

# ── IBT 8-Year Proprietary Benchmarks (183 students, 6 schools) ─────────────
IBT_BENCHMARKS = {
    "overall":         43.3,
    "Mathematics":     39.1,
    "Physics":         39.9,
    "Chemistry":       49.4,
    "Biology":         44.7,
    "English Grammar": 43.3,
    "Literature":      43.3,
    "Economics":       43.3,
    "French":          43.3,
}
IBT_WASSCE_TARGET = 50.0   # B- / minimum WASSCE readiness
IBT_AT_RISK_LINE  = 37.5   # C- / intervention trigger

SUBJECT_SHEETS = [
    "Mathematics", "English Grammar", "Literature",
    "Physics", "Chemistry", "Biology", "Economics", "French"
]

# ── Column Index Map for IBT Grade Tracker Subject Sheets ───────────────────
# Row layout (0-indexed within each row tuple):
#  Col 0:      Student name
#  Col 1:      spacer
#  Cols 2-16:  Sem1 HW 1-15
#  Cols 17-31: Sem1 Quiz 1-15
#  Cols 32-34: Sem1 Test 1-3
#  Col 35:     Sem1 Weighted Avg
#  Col 36:     spacer
#  Cols 37-51: Sem2 HW 1-15
#  Cols 52-66: Sem2 Quiz 1-15
#  Cols 67-69: Sem2 Test 4-6
#  Col 70:     Sem2 Weighted Avg

S1_HW    = list(range(2,  17))
S1_QZ    = list(range(17, 32))
S1_TEST  = list(range(32, 35))
S1_AVG   = 35
S2_HW    = list(range(37, 52))
S2_QZ    = list(range(52, 67))
S2_TEST  = list(range(67, 70))
S2_AVG   = 70


# ── Helper Functions ─────────────────────────────────────────────────────────

def _safe_num(v):
    """Convert a cell value to float in [0,100] or None."""
    try:
        f = float(v)
        return round(f, 1) if 0 <= f <= 100 else None
    except (TypeError, ValueError):
        return None


def _wavg(hw, qz, tests):
    """Weighted average: HW 10%, Quiz 20%, Period Tests 70%."""
    parts, weights = [], []
    if hw:    parts.append(sum(hw)/len(hw));    weights.append(0.10)
    if qz:    parts.append(sum(qz)/len(qz));    weights.append(0.20)
    if tests: parts.append(sum(tests)/len(tests)); weights.append(0.70)
    if not parts:
        return None
    total_w = sum(weights)
    return round(sum(p*w for p,w in zip(parts,weights)) / total_w, 1)


def letter_grade(score):
    if score is None: return "N/A"
    if score >= 80:   return "A"
    if score >= 70:   return "B+"
    if score >= 65:   return "B"
    if score >= 60:   return "B-"
    if score >= 50:   return "C"
    if score >= 44:   return "C"
    if score >= 37:   return "C-"
    if score >= 25:   return "D"
    return "F"


def ibt_status(score, subject="overall"):
    bench = IBT_BENCHMARKS.get(subject, IBT_BENCHMARKS["overall"])
    if score is None:
        return "No data", "#8899BB"
    if score >= bench + 10:
        return "Above IBT avg ✅", C_GREEN
    if score >= bench:
        return "At IBT avg ✅", "#81C784"
    if score >= bench - 8:
        return "Near IBT avg ⚠️", C_AMBER
    return "Below IBT avg 🔴", C_RED


# ── Excel Loader ─────────────────────────────────────────────────────────────

def load_excel_data(file_bytes):
    """
    Load IBT Grade Tracker Excel.
    Returns (roster, subject_data, error_msg).

    roster: list of dicts {name, student_id, grade, gender}
    subject_data: {subject: {student_name: {
        hw1, qz1, test1, s1_avg,
        hw2, qz2, test2, s2_avg,
        overall_avg
    }}}
    """
    try:
        import openpyxl
    except ImportError:
        return None, None, "openpyxl not installed"

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        return None, None, f"Cannot open Excel file: {e}"

    # ── Roster ──────────────────────────────────────────────────────────────
    roster = []
    if "Roster" in wb.sheetnames:
        ws = wb["Roster"]
        # Row 1-2: titles; Row 3: column headers; Rows 4+: students
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row or not row[0]:
                continue
            active = str(row[8]).strip().upper() if len(row) > 8 and row[8] else "YES"
            if active == "NO":
                continue
            roster.append({
                "name":       str(row[0]).strip(),
                "student_id": str(row[1]).strip() if row[1] else "",
                "grade":      str(row[2]).strip() if row[2] else "",
                "gender":     str(row[3]).strip() if row[3] else "",
            })

    # ── Subject Sheets ───────────────────────────────────────────────────────
    subject_data = {}
    for subj in SUBJECT_SHEETS:
        if subj not in wb.sheetnames:
            continue
        ws = wb[subj]
        subject_data[subj] = {}

        # Data starts at row 9 (1-indexed); rows 1-8 are headers
        for row in ws.iter_rows(min_row=9, values_only=True):
            if not row or not row[0] or str(row[0]).strip() in ("", "GRADES"):
                continue
            name = str(row[0]).strip()

            hw1    = [_safe_num(row[i]) for i in S1_HW   if i < len(row) and _safe_num(row[i]) is not None]
            qz1    = [_safe_num(row[i]) for i in S1_QZ   if i < len(row) and _safe_num(row[i]) is not None]
            test1  = [_safe_num(row[i]) for i in S1_TEST  if i < len(row) and _safe_num(row[i]) is not None]
            s1_avg = _safe_num(row[S1_AVG]) if S1_AVG < len(row) else None
            if s1_avg is None:
                s1_avg = _wavg(hw1, qz1, test1)

            hw2    = [_safe_num(row[i]) for i in S2_HW   if i < len(row) and _safe_num(row[i]) is not None]
            qz2    = [_safe_num(row[i]) for i in S2_QZ   if i < len(row) and _safe_num(row[i]) is not None]
            test2  = [_safe_num(row[i]) for i in S2_TEST  if i < len(row) and _safe_num(row[i]) is not None]
            s2_avg = _safe_num(row[S2_AVG]) if S2_AVG < len(row) else None
            if s2_avg is None:
                s2_avg = _wavg(hw2, qz2, test2)

            overall = None
            if s1_avg is not None and s2_avg is not None:
                overall = round((s1_avg + s2_avg) / 2, 1)
            elif s1_avg is not None:
                overall = s1_avg
            elif s2_avg is not None:
                overall = s2_avg

            subject_data[subj][name] = {
                "hw1": hw1, "qz1": qz1, "test1": test1, "s1_avg": s1_avg,
                "hw2": hw2, "qz2": qz2, "test2": test2, "s2_avg": s2_avg,
                "overall_avg": overall,
            }

    return roster, subject_data, None


# ── Main Render Entry Point ──────────────────────────────────────────────────

def render_academic_report_from_excel():
    """
    Primary render function — call this inside Tab 5 of app.py.
    Handles data source selection, loading, and dispatch to sub-renderers.
    """
    import datetime as _ardt

    school_label = st.session_state.get("_classroom_label",
                   st.session_state.get("_school_v", "My School"))
    grade_en     = st.session_state.get("grade_en", "")

    # ── Header Banner ────────────────────────────────────────────────────────
    st.markdown(f"""
<div style="background:linear-gradient(135deg,#0D1B2A,#0D3B8C);border-radius:14px;
  padding:20px 24px 16px;margin-bottom:16px;border:1px solid #1E3A6A">
  <div style="display:flex;align-items:center;gap:14px;margin-bottom:8px">
    <div style="font-size:2rem">🎓</div>
    <div>
      <div style="color:{C_GOLD};font-size:1.1rem;font-weight:800;letter-spacing:.5px">
        INSTITUTE OF BASIC TECHNOLOGY</div>
      <div style="color:#ffffff;font-size:.92rem;font-weight:600">
        Teacher Pehpeh — Academic Performance Report</div>
    </div>
  </div>
  <div style="border-top:1px solid rgba(212,168,67,.3);padding-top:10px;
    display:flex;gap:32px;flex-wrap:wrap">
    <div><span style="color:#8899aa;font-size:.8rem">SCHOOL</span>
         <br><span style="color:#fff;font-weight:700">{school_label}</span></div>
    <div><span style="color:#8899aa;font-size:.8rem">GRADE LEVEL</span>
         <br><span style="color:#fff;font-weight:700">{grade_en or "—"}</span></div>
    <div><span style="color:#8899aa;font-size:.8rem">GENERATED</span>
         <br><span style="color:#fff;font-weight:700">
           {_ardt.datetime.now().strftime("%B %d, %Y")}</span></div>
    <div><span style="color:#8899aa;font-size:.8rem">DATA SOURCE</span>
         <br><span style="color:#fff;font-weight:700">IBT Grade Tracker v2</span></div>
  </div>
</div>""", unsafe_allow_html=True)

    # ── Data Source Selector ─────────────────────────────────────────────────
    dsrc = st.radio(
        "**Data Source**",
        ["📊 Upload IBT Grade Tracker Excel", "📝 Use Manually-Entered Grades"],
        horizontal=True, key="ar_data_source_v2"
    )

    # Show "switch back" button when Excel data is cached and user picks manual mode
    if st.session_state.get("_ar_subject_data") and dsrc == "📝 Use Manually-Entered Grades":
        cc1, cc2 = st.columns([3,1])
        with cc2:
            if st.button("🗑️ Clear Excel grades", key="ar_clear_excel",
                         help="Removes loaded Excel grades — re-enables camera/manual entry in Students tab"):
                st.session_state.pop("_ar_subject_data", None)
                st.session_state.pop("_ar_roster", None)
                st.rerun()

    if dsrc == "📊 Upload IBT Grade Tracker Excel":
        _render_excel_path(school_label, grade_en)
    else:
        _render_session_path(school_label, grade_en)


# ── Excel Path ───────────────────────────────────────────────────────────────

def _render_excel_path(school_label, grade_en):
    f = st.file_uploader(
        "Upload IBT_Student_Grade_Tracker.xlsx",
        type=["xlsx", "xls"],
        key="ar_excel_v2",
        help="Upload the IBT Grade Tracker. Grades in each subject sheet will be read automatically."
    )

    if not f:
        st.markdown(f"""
<div style="background:rgba(212,168,67,.05);border:1px solid rgba(212,168,67,.2);
  border-radius:10px;padding:16px 18px;margin-top:6px">
  <p style="color:{C_GOLD};font-weight:700;margin-bottom:8px">
    📋 IBT Grade Tracker — Expected Format</p>
  <p style="color:#8899BB;font-size:.85rem;margin-bottom:6px">
    The Excel should contain these sheets:</p>
  <ul style="color:#D0D8E8;font-size:.84rem;margin-left:16px;line-height:1.9">
    <li><strong>Roster</strong> — student name, ID, grade, gender (row 4+)</li>
    <li><strong>Mathematics, Physics, Chemistry, Biology, English Grammar, Literature,
        Economics, French</strong> — one row per student with HW/Quiz/Test scores starting at row 9</li>
  </ul>
  <p style="color:#556;font-size:.78rem;margin-top:8px">
    Score weights: Homework 10% · Quizzes 20% · Period Tests 70%</p>
</div>""", unsafe_allow_html=True)
        return

    file_bytes = f.read()
    with st.spinner("📊 Parsing grade tracker…"):
        roster, subject_data, err = load_excel_data(file_bytes)

    if err:
        st.error(f"❌ {err}")
        return

    if not roster:
        st.warning("⚠️ No active students found in the Roster sheet.")
        return

    # ── Cache in session state so IBT What-If tab can read it ────────────────
    st.session_state["_ar_roster"]       = roster
    st.session_state["_ar_subject_data"] = subject_data

    has_grades = subject_data and any(len(v) > 0 for v in subject_data.values())

    if not has_grades:
        st.info(f"📭 Found {len(roster)} students in Roster but no subject grades yet. "
                "Fill scores in the subject sheets and re-upload.")
        st.markdown(
            f'<p style="color:{C_GOLD};font-weight:600;margin-top:8px">'
            f'{len(roster)} students in Roster:</p>', unsafe_allow_html=True)
        for s in roster:
            st.markdown(
                f'<span style="color:#D0D8E8;font-size:.88rem">'
                f'• {s["name"]} ({s["grade"]})</span><br>', unsafe_allow_html=True)
        return

    _render_full_report(roster, subject_data, school_label, grade_en)


# ── Full Report Renderer ─────────────────────────────────────────────────────

def _render_full_report(roster, subject_data, school_label, grade_en):
    """
    Main report: class summary metrics → student selector → individual deep-dive.
    """
    try:
        import pandas as pd
    except ImportError:
        st.error("pandas required for report generation")
        return

    all_subjects = [s for s in SUBJECT_SHEETS if s in subject_data and subject_data[s]]

    # ── Class-Level Metrics ──────────────────────────────────────────────────
    student_overall_avgs = {}
    for s in roster:
        scores = [subject_data[subj][s["name"]]["overall_avg"]
                  for subj in all_subjects
                  if s["name"] in subject_data.get(subj, {})
                  and subject_data[subj][s["name"]]["overall_avg"] is not None]
        if scores:
            student_overall_avgs[s["name"]] = round(sum(scores)/len(scores), 1)

    class_avg   = round(sum(student_overall_avgs.values()) / max(1, len(student_overall_avgs)), 1)
    below50     = sum(1 for v in student_overall_avgs.values() if v < 50)
    ibt_bench   = IBT_BENCHMARKS["overall"]
    gap_to_bench = round(class_avg - ibt_bench, 1)

    m1, m2, m3, m4 = st.columns(4)
    with m1: st.metric("Class Average", f"{class_avg:.1f}/100",
                        delta=f"{gap_to_bench:+.1f} vs IBT avg")
    with m2: st.metric("Students Tracked", len(roster))
    with m3: st.metric("Subjects with Data", len(all_subjects))
    with m4: st.metric("Need Intervention", f"{below50}",
                        delta=f"{below50/max(1,len(roster))*100:.0f}% of class",
                        delta_color="inverse")

    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

    # ── View Toggle ──────────────────────────────────────────────────────────
    view = st.radio("**View**",
                    ["🏫 Class Overview", "👤 Student Deep-Dive"],
                    horizontal=True, key="ar_view_v2")

    if view == "🏫 Class Overview":
        _render_class_overview(roster, subject_data, all_subjects,
                               student_overall_avgs, class_avg, pd)
    else:
        _render_student_deep_dive(roster, subject_data, all_subjects,
                                  student_overall_avgs, pd)


# ── Class Overview ───────────────────────────────────────────────────────────

def _render_class_overview(roster, subject_data, all_subjects,
                            student_overall_avgs, class_avg, pd):
    try:
        import altair as alt
    except ImportError:
        alt = None

    # Summary table
    st.markdown(
        f'<div style="color:{C_GOLD};font-weight:700;font-size:.9rem;margin-bottom:6px">'
        '📋 Student Performance Summary</div>', unsafe_allow_html=True)

    rows = []
    for s in roster:
        name  = s["name"]
        ov    = student_overall_avgs.get(name)
        bench = IBT_BENCHMARKS["overall"]
        gap   = f"{ov - bench:+.1f}" if ov is not None else "—"
        gl    = letter_grade(ov)
        status_label, _ = ibt_status(ov)
        row = {
            "Student":   name,
            "Grade":     s["grade"],
            "Overall":   f"{ov:.0f}/100" if ov is not None else "—",
            "IBT Gap":   gap,
            "Grade":     gl,
            "Status":    status_label,
        }
        for subj in all_subjects:
            sc = subject_data.get(subj, {}).get(name, {}).get("overall_avg")
            row[subj[:7]] = f"{sc:.0f}" if sc is not None else "—"
        rows.append(row)

    df = pd.DataFrame(rows)
    st.dataframe(df, use_container_width=True, hide_index=True)

    # Subject vs IBT Benchmark bar chart
    bench_rows = []
    for subj in all_subjects:
        scores = [subject_data[subj][s["name"]]["overall_avg"]
                  for s in roster
                  if s["name"] in subject_data.get(subj, {})
                  and subject_data[subj][s["name"]]["overall_avg"] is not None]
        if scores:
            avg   = round(sum(scores)/len(scores), 1)
            bench = IBT_BENCHMARKS.get(subj, IBT_BENCHMARKS["overall"])
            bench_rows.append({
                "Subject":       subj,
                "Class Avg":     avg,
                "IBT Benchmark": bench,
            })

    if bench_rows and alt:
        st.markdown(
            f'<div style="color:{C_GOLD};font-weight:700;font-size:.9rem;margin:12px 0 6px">'
            '📊 Class Average vs IBT Benchmarks by Subject</div>', unsafe_allow_html=True)
        bdf = pd.DataFrame(bench_rows)
        bdf_m = bdf.melt("Subject", var_name="Series", value_name="Score")
        chart = (
            alt.Chart(bdf_m)
            .mark_bar(opacity=0.88, cornerRadiusTopLeft=3, cornerRadiusTopRight=3)
            .encode(
                x=alt.X("Subject:N", title="", axis=alt.Axis(labelAngle=-30)),
                y=alt.Y("Score:Q", scale=alt.Scale(domain=[0, 100]), title="Score /100"),
                color=alt.Color("Series:N", scale=alt.Scale(
                    domain=["Class Avg", "IBT Benchmark"],
                    range=[C_GOLD, C_BLUE])),
                tooltip=["Subject", "Series", alt.Tooltip("Score:Q", format=".1f")]
            )
            .properties(height=240,
                        title="Class Average vs IBT 8-Year Research Benchmark")
        )
        st.altair_chart(chart, use_container_width=True)

    # At-risk list
    at_risk = [(s["name"], student_overall_avgs[s["name"]])
               for s in roster
               if s["name"] in student_overall_avgs
               and student_overall_avgs[s["name"]] < 50]
    if at_risk:
        at_risk.sort(key=lambda x: x[1])
        items = "".join(
            f'<div style="color:#F0D5D5;font-size:.84rem;margin-top:3px">'
            f'• <strong>{n}</strong>: {sc:.0f}/100 avg '
            f'({sc - IBT_BENCHMARKS["overall"]:+.1f} vs IBT)</div>'
            for n, sc in at_risk
        )
        st.markdown(
            f'<div style="background:rgba(239,83,80,.1);border:2px solid {C_RED};'
            f'border-radius:10px;padding:12px 16px;margin:10px 0">'
            f'<strong style="color:{C_RED}">⚠️ Students Needing Intervention '
            f'({len(at_risk)})</strong>{items}</div>',
            unsafe_allow_html=True)


# ── Student Deep-Dive ─────────────────────────────────────────────────────────

def _render_student_deep_dive(roster, subject_data, all_subjects,
                               student_overall_avgs, pd):
    try:
        import altair as alt
    except ImportError:
        alt = None

    student_names = [s["name"] for s in roster]
    sel = st.selectbox("Select student:", student_names, key="ar_stu_sel_v2")
    stu_info = next((s for s in roster if s["name"] == sel), {})

    ov = student_overall_avgs.get(sel)
    bench = IBT_BENCHMARKS["overall"]
    status_label, status_color = ibt_status(ov)
    grade_ltr = letter_grade(ov)

    # Student header card
    st.markdown(f"""
<div style="background:rgba(13,59,140,.15);border:1px solid #1E3A6A;
  border-radius:12px;padding:14px 20px;margin-bottom:14px;
  display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:10px">
  <div>
    <div style="color:{C_GOLD};font-size:1.05rem;font-weight:800">{sel}</div>
    <div style="color:#8899BB;font-size:.82rem">
      {stu_info.get("grade","")} &nbsp;·&nbsp; ID: {stu_info.get("student_id","—")}
      &nbsp;·&nbsp; {stu_info.get("gender","—")}
    </div>
  </div>
  <div style="text-align:right">
    <div style="font-size:1.8rem;font-weight:800;color:#fff">
      {f"{ov:.0f}" if ov is not None else "—"}<span style="font-size:.9rem;color:#8899BB">/100</span>
    </div>
    <div style="color:{status_color};font-size:.8rem">{status_label}</div>
    <div style="color:#556;font-size:.75rem">Grade {grade_ltr} · IBT bench: {bench}</div>
  </div>
</div>""", unsafe_allow_html=True)

    # Subject score cards
    student_subjects = {
        subj: subject_data[subj][sel]
        for subj in all_subjects
        if sel in subject_data.get(subj, {})
    }

    if not student_subjects:
        st.info(f"No grade data found for {sel} in any subject sheet.")
        return

    cols_per_row = 4
    subj_list = list(student_subjects.items())
    for batch_start in range(0, len(subj_list), cols_per_row):
        batch = subj_list[batch_start:batch_start + cols_per_row]
        cols  = st.columns(len(batch))
        for col, (subj, data) in zip(cols, batch):
            s1  = data.get("s1_avg")
            s2  = data.get("s2_avg")
            ov2 = data.get("overall_avg")
            b2  = IBT_BENCHMARKS.get(subj, bench)
            slabel, scolor = ibt_status(ov2, subj)
            gl2 = letter_grade(ov2)
            s1s2 = (f'<div style="color:#8899BB;font-size:.72rem;margin-top:3px">'
                    f'S1: {s1:.0f} &nbsp;·&nbsp; S2: {s2:.0f}</div>'
                    if s1 is not None and s2 is not None else "")
            col.markdown(f"""
<div style="background:#0F1E38;border:1px solid #1E3A6A;border-radius:10px;
  padding:12px;text-align:center;margin-bottom:6px">
  <div style="color:{C_GOLD};font-weight:700;font-size:.8rem;margin-bottom:3px">
    {subj[:12]}</div>
  <div style="font-size:1.55rem;font-weight:800;color:#fff;line-height:1.1">
    {f"{ov2:.0f}" if ov2 is not None else "—"}</div>
  <div style="color:#8899BB;font-size:.73rem">/100 · Grade {gl2}</div>
  <div style="color:{scolor};font-size:.72rem;margin-top:3px">{slabel}</div>
  <div style="color:#445;font-size:.69rem">bench: {b2}</div>
  {s1s2}
</div>""", unsafe_allow_html=True)

    # Semester progression line chart (Altair with hover)
    chart_rows = []
    for subj, data in student_subjects.items():
        bench_val = IBT_BENCHMARKS.get(subj, bench)
        s1 = data.get("s1_avg")
        s2 = data.get("s2_avg")
        if s1 is not None:
            chart_rows.append({
                "Subject": subj, "Period": "Semester 1", "Score": s1,
                "IBT Benchmark": bench_val,
                "Gap": round(s1 - bench_val, 1),
            })
        if s2 is not None:
            chart_rows.append({
                "Subject": subj, "Period": "Semester 2", "Score": s2,
                "IBT Benchmark": bench_val,
                "Gap": round(s2 - bench_val, 1),
            })

    if len(chart_rows) >= 2 and alt:
        st.markdown(
            f'<div style="color:{C_GOLD};font-weight:700;font-size:.9rem;margin:12px 0 6px">'
            '📈 Semester Progression by Subject (hover for details)</div>',
            unsafe_allow_html=True)
        cdf = pd.DataFrame(chart_rows)

        student_line = (
            alt.Chart(cdf)
            .mark_line(point=alt.OverlayMarkDef(size=70, filled=True), strokeWidth=2.5)
            .encode(
                x=alt.X("Period:O", title=""),
                y=alt.Y("Score:Q", scale=alt.Scale(domain=[0, 100]), title="Score /100"),
                color=alt.Color("Subject:N"),
                tooltip=[
                    "Subject", "Period",
                    alt.Tooltip("Score:Q",         format=".1f", title="Score"),
                    alt.Tooltip("IBT Benchmark:Q", format=".1f", title="IBT Bench"),
                    alt.Tooltip("Gap:Q",           format="+.1f", title="Gap to IBT"),
                ]
            )
        )

        # IBT benchmark rule per subject (use mean as a flat line per subject)
        bench_rule = (
            alt.Chart(cdf)
            .mark_rule(strokeDash=[6, 4], opacity=0.4, strokeWidth=1.5)
            .encode(
                y=alt.Y("mean(IBT Benchmark):Q"),
                color=alt.value(C_BLUE),
                tooltip=[alt.Tooltip("mean(IBT Benchmark):Q", format=".1f", title="IBT Avg")]
            )
        )

        st.altair_chart(
            (student_line + bench_rule).properties(
                height=260,
                title=f"{sel} — Score Progression (blue dashed = IBT avg per subject)"
            ),
            use_container_width=True
        )

    # Category breakdown table
    cat_rows = []
    for subj, data in student_subjects.items():
        hw1    = data.get("hw1",    [])
        qz1    = data.get("qz1",    [])
        test1  = data.get("test1",  [])
        hw2    = data.get("hw2",    [])
        qz2    = data.get("qz2",    [])
        test2  = data.get("test2",  [])
        s1_avg = data.get("s1_avg")
        s2_avg = data.get("s2_avg")
        ov2    = data.get("overall_avg")

        if not (hw1 or qz1 or test1 or hw2 or qz2 or test2):
            continue

        def _fmt_list(lst):
            return f"{sum(lst)/len(lst):.0f}" if lst else "—"

        cat_rows.append({
            "Subject":         subj,
            "HW Avg S1 (10%)": _fmt_list(hw1),
            "Quiz Avg S1 (20%)": _fmt_list(qz1),
            "Test Avg S1 (70%)": _fmt_list(test1),
            "Sem 1 Avg":       f"{s1_avg:.0f}" if s1_avg else "—",
            "HW Avg S2 (10%)": _fmt_list(hw2),
            "Quiz Avg S2 (20%)": _fmt_list(qz2),
            "Test Avg S2 (70%)": _fmt_list(test2),
            "Sem 2 Avg":       f"{s2_avg:.0f}" if s2_avg else "—",
            "Overall":         f"{ov2:.0f}" if ov2 else "—",
        })

    if cat_rows:
        st.markdown(
            f'<div style="color:{C_GOLD};font-weight:700;font-size:.9rem;margin:12px 0 6px">'
            '📋 Homework / Quiz / Test Breakdown</div>', unsafe_allow_html=True)
        st.dataframe(pd.DataFrame(cat_rows), use_container_width=True, hide_index=True)


# ── Session-State Fallback Path ──────────────────────────────────────────────

def _render_session_path(school_label, grade_en):
    """Show grade history from manually entered scores (original behavior)."""
    try:
        import pandas as pd
        import altair as alt
        import datetime as _ardt
    except ImportError:
        st.error("pandas required")
        return

    _gh = st.session_state.get("grade_history", [])
    if not _gh:
        st.info("📭 No grade data yet. Enter grades in the Students tab, "
                "or switch to 'Upload IBT Grade Tracker Excel' above.")
        return

    _df = pd.DataFrame(_gh)
    _df["date"] = pd.to_datetime(_df["date"])
    _df_sorted  = _df.sort_values("date")

    # Metrics
    avg_all    = _df["score"].mean()
    n_students = _df["student"].nunique()
    latest3    = _df_sorted.tail(max(1, len(_df)//3))["score"].mean()
    earliest3  = _df_sorted.head(max(1, len(_df)//3))["score"].mean()
    trend_d    = latest3 - earliest3
    trend_icon = "📈" if trend_d > 2 else ("📉" if trend_d < -2 else "➡️")
    below50    = _df[_df["score"] < 50]["student"].nunique()

    m1, m2, m3, m4 = st.columns(4)
    with m1: st.metric("Class Average", f"{avg_all:.1f}/100")
    with m2: st.metric("Trend", f"{trend_icon} {abs(trend_d):.1f} pts",
                        delta=f"{trend_d:+.1f}")
    with m3: st.metric("Total Records", len(_df))
    with m4: st.metric("Need Intervention", f"{below50}",
                        delta_color="inverse",
                        delta=f"{below50/max(1,n_students)*100:.0f}%")

    # Per-student table
    st.markdown(
        f'<div style="color:{C_GOLD};font-weight:700;font-size:.9rem;margin:10px 0 6px">'
        '📋 Student Performance Summary</div>', unsafe_allow_html=True)
    rows = []
    for sname in sorted(_df["student"].unique()):
        sd    = _df[_df["student"] == sname].sort_values("date")
        savg  = sd["score"].mean()
        slast = sd["score"].iloc[-1]
        std_d = sd["score"].iloc[-1] - sd["score"].iloc[0] if len(sd) > 1 else 0
        str_i = "📈" if std_d > 2 else ("📉" if std_d < -2 else "➡️")
        gap   = f"{savg - IBT_BENCHMARKS['overall']:+.1f}"
        rows.append({
            "Student":    sname,
            "Avg Score":  f"{savg:.0f}/100",
            "Latest":     f"{slast}/100",
            "Trend":      str_i,
            "IBT Gap":    gap,
            "Subjects":   ", ".join(sorted(sd["subject"].unique())),
            "Status":     "🔴 Intervention" if savg < 50 else ("🟡 Monitor" if savg < 65 else "🟢 On Track"),
        })
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    # Line chart with hover
    try:
        chart_df = _df_sorted[["date", "student", "score", "subject"]].copy()
        chart_df["date"] = chart_df["date"].dt.date
        ac = (
            alt.Chart(chart_df)
            .mark_line(point=True)
            .encode(
                x=alt.X("date:T", title="Date"),
                y=alt.Y("score:Q", title="Score", scale=alt.Scale(domain=[0, 100])),
                color="student:N",
                tooltip=["student", "score", "subject", "date"]
            )
            .properties(height=260, title="Student Score Progression")
        )
        # IBT benchmark rule
        bench_df = pd.DataFrame({"y": [IBT_BENCHMARKS["overall"]]})
        bench_rule = (
            alt.Chart(bench_df)
            .mark_rule(strokeDash=[8, 4], color=C_BLUE, opacity=0.6)
            .encode(y="y:Q", tooltip=[alt.Tooltip("y:Q", title="IBT Avg Benchmark")])
        )
        st.altair_chart((ac + bench_rule), use_container_width=True)
    except Exception:
        pass
