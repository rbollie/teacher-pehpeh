"""
grade_tracker.py  —  IBT Grade Tracker Integration for Teacher Pehpeh
=====================================================================
Drop this into your Teacher Pehpeh project root.

Usage in Streamlit:
    from grade_tracker import GradeTracker, render_grade_report

Key functions:
    tracker = GradeTracker(uploaded_file)
    tracker.summary()          → dict of per-subject averages + status
    render_grade_report(tracker)  → full Streamlit report UI
"""

import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime

# ── Constants ──────────────────────────────────────────────────────────────────
SUBJECTS = [
    "Mathematics", "English Grammar", "Literature",
    "Physics", "Chemistry", "Biology",
    "Economics", "French"
]

W_HW, W_QUIZ, W_TEST = 0.10, 0.20, 0.70

# Column layout in each subject sheet (1-indexed)
# Semester 1: HW C–Q (3–17), Quiz R–AF (18–32), Test AG–AI (33–35), Avg AJ (36)
# Semester 2: HW AL–AZ (38–52), Quiz BA–BO (53–67), Test BP–BR (68–70), Avg BS (71)
SEM1 = dict(hw=(3,17),  quiz=(18,32), test=(33,35), avg=36, grade_row=8)
SEM2 = dict(hw=(38,52), quiz=(53,67), test=(68,70), avg=71, grade_row=8)

# Intervention thresholds
THRESHOLDS = [
    (85, "⭐ Excelling",        "#1B4F8A", "#D6E4F0", "Consider enrichment/advanced challenges"),
    (75, "🟢 Meeting Standard", "#27AE60", "#D5F0E4", "On track — maintain current effort"),
    (65, "🟡 On Track",         "#F39C12", "#FEF9E7", "Monitor weekly — minor support may help"),
    (50, "🟠 Needs Monitoring", "#E67E22", "#FEF3D0", "Schedule tutoring / extra support sessions"),
    (0,  "🔴 At Risk",          "#C0392B", "#FADBD8", "URGENT: Immediate intervention required"),
]

def classify(score):
    """Return (label, color, bg_color, action) for a given score."""
    if score is None or np.isnan(score):
        return "⬜ No Data", "#888888", "#F2F2F2", "Awaiting grades"
    for threshold, label, color, bg, action in THRESHOLDS:
        if score >= threshold:
            return label, color, bg, action
    return "🔴 At Risk", "#C0392B", "#FADBD8", "URGENT: Immediate intervention required"


def _col_range_vals(ws, col_start, col_end, row):
    """Extract numeric values from a row/column range, ignoring blanks."""
    vals = []
    for c in range(col_start, col_end + 1):
        v = ws.cell(row=row, column=c).value
        if v is not None and v != "" and str(v).strip() != "":
            try:
                vals.append(float(v))
            except (ValueError, TypeError):
                pass
    return vals


def _weighted_avg(hw_vals, qz_vals, test_vals):
    """Compute weighted average using only components that have data."""
    total_w, total_score = 0.0, 0.0
    for vals, w in [(hw_vals, W_HW), (qz_vals, W_QUIZ), (test_vals, W_TEST)]:
        if vals:
            total_score += np.mean(vals) * w
            total_w += w
    return (total_score / total_w) if total_w > 0 else None


def _running_avg(vals):
    """Return list of cumulative running averages."""
    result, running = [], []
    for v in vals:
        if v is not None:
            running.append(v)
        result.append(np.mean(running) if running else None)
    return result


class GradeTracker:
    """Parses an IBT Grade Tracker .xlsx file and exposes analysis methods."""

    def __init__(self, file_source):
        """
        file_source: Streamlit UploadedFile, BytesIO, or file path str.
        """
        if isinstance(file_source, str):
            self.wb = load_workbook(file_source, data_only=True)
        else:
            content = file_source.read() if hasattr(file_source, 'read') else file_source
            self.wb = load_workbook(BytesIO(content), data_only=True)

        self._parse_student_info()
        self._parse_all_subjects()

    def _parse_student_info(self):
        try:
            ws = self.wb["Student_Info"]
            self.student_name = ws['B5'].value or "Unknown Student"
            self.student_id   = ws['B6'].value or "—"
            self.grade_level  = ws['B7'].value or "—"
            self.academic_year = ws['B8'].value or "—"
            self.school       = ws['B10'].value or "IBT"
        except Exception:
            self.student_name  = "Unknown Student"
            self.student_id    = "—"
            self.grade_level   = "—"
            self.academic_year = "—"
            self.school        = "IBT"

    def _parse_subject(self, subject_name):
        """Return dict with parsed grade data for one subject."""
        try:
            ws = self.wb[subject_name[:31]]
        except KeyError:
            return None

        row = SEM1['grade_row']

        result = {}
        for sem_key, cfg in [('sem1', SEM1), ('sem2', SEM2)]:
            hw_vals   = _col_range_vals(ws, cfg['hw'][0],   cfg['hw'][1],   row)
            qz_vals   = _col_range_vals(ws, cfg['quiz'][0], cfg['quiz'][1], row)
            test_vals = _col_range_vals(ws, cfg['test'][0], cfg['test'][1], row)
            wtd       = _weighted_avg(hw_vals, qz_vals, test_vals)

            result[sem_key] = {
                'hw_vals':   hw_vals,
                'qz_vals':   qz_vals,
                'test_vals': test_vals,
                'hw_avg':    np.mean(hw_vals)   if hw_vals   else None,
                'qz_avg':    np.mean(qz_vals)   if qz_vals   else None,
                'test_avg':  np.mean(test_vals) if test_vals else None,
                'wtd_avg':   wtd,
                'hw_running':   _running_avg(hw_vals),
                'qz_running':   _running_avg(qz_vals),
                'test_running': _running_avg(test_vals),
                'n_hw':   len(hw_vals),
                'n_qz':   len(qz_vals),
                'n_test': len(test_vals),
            }

        # Year average (average of available semester avgs)
        avgs = [v for v in [result['sem1']['wtd_avg'], result['sem2']['wtd_avg']] if v is not None]
        result['year_avg'] = np.mean(avgs) if avgs else None
        result['subject']  = subject_name
        return result

    def _parse_all_subjects(self):
        self.data = {}
        for subj in SUBJECTS:
            d = self._parse_subject(subj)
            if d:
                self.data[subj] = d

    def summary(self):
        """Return list of dicts: one per subject with averages and intervention status."""
        rows = []
        for subj, d in self.data.items():
            s1 = d['sem1']['wtd_avg']
            s2 = d['sem2']['wtd_avg']
            yr = d['year_avg']
            label, color, bg, action = classify(yr if yr is not None else s1)
            rows.append({
                'subject':      subj,
                'sem1_avg':     s1,
                'sem2_avg':     s2,
                'year_avg':     yr,
                'status_label': label,
                'status_color': color,
                'status_bg':    bg,
                'action':       action,
                'sem1_n_hw':    d['sem1']['n_hw'],
                'sem1_n_qz':    d['sem1']['n_qz'],
                'sem1_n_test':  d['sem1']['n_test'],
            })
        return rows

    def at_risk_subjects(self):
        return [r for r in self.summary() if "At Risk" in r['status_label'] or "Monitoring" in r['status_label']]

    def has_data(self):
        return any(d['year_avg'] is not None or d['sem1']['wtd_avg'] is not None
                   for d in self.data.values())


# ── Streamlit UI ───────────────────────────────────────────────────────────────

def render_grade_report(tracker: GradeTracker):
    """Full Streamlit grade/intervention report. Call this from Teacher Pehpeh."""
    import streamlit as st

    # ── Header ─────────────────────────────────────────────────────────────────
    st.markdown(f"""
    <div style='background:#1B4F8A;border-radius:8px;padding:16px 20px;margin-bottom:12px'>
      <h2 style='color:white;margin:0;font-size:20px'>📊 IBT Academic Progress Report</h2>
      <p style='color:#D6E4F0;margin:4px 0 0'>
        <b>{tracker.student_name}</b> &nbsp;|&nbsp; ID: {tracker.student_id}
        &nbsp;|&nbsp; {tracker.grade_level} &nbsp;|&nbsp; {tracker.academic_year}
      </p>
      <p style='color:#A8C8E8;margin:2px 0 0;font-size:12px'>
        Generated: {datetime.now().strftime("%d %b %Y, %I:%M %p")}
      </p>
    </div>
    """, unsafe_allow_html=True)

    if not tracker.has_data():
        st.warning("⚠️ No grade data found in this file. Please fill in grades on the subject sheets.")
        return

    summary = tracker.summary()

    # ── At-Risk Alert Banner ────────────────────────────────────────────────────
    at_risk = [r for r in summary if "At Risk" in r['status_label']]
    monitor = [r for r in summary if "Monitoring" in r['status_label']]

    if at_risk:
        subjects_str = ", ".join(r['subject'] for r in at_risk)
        st.error(f"🚨 **URGENT INTERVENTION REQUIRED** — {subjects_str}\n\nThese subjects are below 50%. Immediate action needed.")
    if monitor:
        subjects_str = ", ".join(r['subject'] for r in monitor)
        st.warning(f"⚠️ **Monitoring Required** — {subjects_str}\n\nSchedule tutoring sessions for these subjects.")

    # ── Dashboard Table ─────────────────────────────────────────────────────────
    st.markdown("### 📋 All-Subject Overview")

    cols = st.columns([2.5, 1.2, 1.2, 1.2, 2.2, 2.8])
    headers = ["Subject", "Sem 1 %", "Sem 2 %", "Year %", "Status", "Recommended Action"]
    for col, h in zip(cols, headers):
        col.markdown(f"**{h}**")
    st.markdown("<hr style='margin:4px 0 8px'>", unsafe_allow_html=True)

    for row in summary:
        cols = st.columns([2.5, 1.2, 1.2, 1.2, 2.2, 2.8])
        cols[0].markdown(f"**{row['subject']}**")

        def fmt(v):
            return f"{v:.1f}" if v is not None else "—"

        cols[1].markdown(fmt(row['sem1_avg']))
        cols[2].markdown(fmt(row['sem2_avg']))
        yr = row['year_avg']
        cols[3].markdown(
            f"<b style='color:{row['status_color']}'>{fmt(yr)}</b>"
            if yr is not None else "—",
            unsafe_allow_html=True
        )
        cols[4].markdown(
            f"<span style='background:{row['status_bg']};color:{row['status_color']};"
            f"padding:2px 8px;border-radius:4px;font-size:13px;font-weight:600'>"
            f"{row['status_label']}</span>",
            unsafe_allow_html=True
        )
        cols[5].markdown(
            f"<span style='font-size:12px;color:#444'>{row['action']}</span>",
            unsafe_allow_html=True
        )

    # ── Detailed Subject Drill-Down ─────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 🔍 Subject Detail  *(moving averages as grades came in)*")

    has_any = any(
        tracker.data[s]['sem1']['n_hw'] + tracker.data[s]['sem1']['n_qz'] + tracker.data[s]['sem1']['n_test'] > 0
        for s in tracker.data
    )

    if has_any:
        subject_choice = st.selectbox(
            "Select subject to inspect:",
            [s for s in SUBJECTS if s in tracker.data],
            key="grade_subject_select"
        )

        if subject_choice:
            d = tracker.data[subject_choice]
            _render_subject_detail(subject_choice, d)

    # ── IBT Intervention Targets ────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 🎯 IBT Intervention Target Summary")

    target_data = {
        "Category": ["At Risk", "Needs Monitoring", "On Track", "Meeting Standard", "Excelling"],
        "Score Range": ["< 50%", "50–64%", "65–74%", "75–84%", "≥ 85%"],
        "This Student": [""] * 5,
        "IBT Action": [
            "Immediate 1-on-1 intervention",
            "Weekly tutoring sessions",
            "Peer study groups",
            "Enrichment exercises",
            "Advanced/leadership track"
        ]
    }

    counts = [0, 0, 0, 0, 0]
    for row in summary:
        y = row['year_avg'] if row['year_avg'] is not None else row['sem1_avg']
        if y is None: continue
        if y < 50:   counts[0] += 1
        elif y < 65: counts[1] += 1
        elif y < 75: counts[2] += 1
        elif y < 85: counts[3] += 1
        else:        counts[4] += 1

    target_data["This Student"] = [
        f"{c} subject(s)" if c > 0 else "—" for c in counts
    ]

    df = pd.DataFrame(target_data)
    st.dataframe(df, use_container_width=True, hide_index=True)

    # ── Download report as CSV ──────────────────────────────────────────────────
    st.markdown("---")
    csv_rows = []
    for row in summary:
        csv_rows.append({
            "Student": tracker.student_name,
            "Student ID": tracker.student_id,
            "Subject": row['subject'],
            "Sem 1 Avg": f"{row['sem1_avg']:.1f}" if row['sem1_avg'] else "",
            "Sem 2 Avg": f"{row['sem2_avg']:.1f}" if row['sem2_avg'] else "",
            "Year Avg": f"{row['year_avg']:.1f}" if row['year_avg'] else "",
            "Status": row['status_label'],
            "Action": row['action'],
            "Report Date": datetime.now().strftime("%Y-%m-%d"),
        })
    df_export = pd.DataFrame(csv_rows)
    csv_bytes = df_export.to_csv(index=False).encode()
    st.download_button(
        label="📥 Download IBT Report (CSV)",
        data=csv_bytes,
        file_name=f"IBT_Report_{tracker.student_name.replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.csv",
        mime="text/csv"
    )


def _render_subject_detail(subject_name, d):
    """Render per-subject moving average detail."""
    import streamlit as st

    tab1, tab2 = st.tabs(["Semester 1", "Semester 2"])

    for tab, sem_key, sem_label in [(tab1, 'sem1', 'Semester 1'), (tab2, 'sem2', 'Semester 2')]:
        with tab:
            s = d[sem_key]
            c1, c2, c3, c4 = st.columns(4)
            def m(v): return f"{v:.1f}" if v is not None else "—"
            c1.metric("HW Average",   m(s['hw_avg']),   f"{s['n_hw']}/15 submitted")
            c2.metric("Quiz Average", m(s['qz_avg']),   f"{s['n_qz']}/15 submitted")
            c3.metric("Test Average", m(s['test_avg']), f"{s['n_test']}/3 written")
            c4.metric("Weighted Avg", m(s['wtd_avg']),  "HW 10% | Qz 20% | Test 70%")

            if s['n_hw'] + s['n_qz'] + s['n_test'] > 0:
                # Build running-avg chart
                chart_data = {}
                if s['hw_vals']:
                    chart_data['HW Running Avg'] = s['hw_running'] + [None] * (15 - len(s['hw_running']))
                if s['qz_vals']:
                    chart_data['Quiz Running Avg'] = s['qz_running'] + [None] * (15 - len(s['qz_running']))
                if chart_data:
                    df_chart = pd.DataFrame(chart_data)
                    df_chart.index = [f"#{i+1}" for i in range(len(df_chart))]
                    st.line_chart(df_chart, use_container_width=True)
            else:
                st.info(f"No grades entered yet for {sem_label}.")


# ── Streamlit Integration Helper ───────────────────────────────────────────────

def grade_tracker_tab():
    """
    Complete self-contained Streamlit tab for grade tracking.
    Add to Teacher Pehpeh by calling this function inside a tab:

        with tab_grades:
            from grade_tracker import grade_tracker_tab
            grade_tracker_tab()
    """
    st.markdown("### 📈 IBT Student Grade Tracker")
    st.markdown(
        "Upload a completed **IBT Grade Tracker** Excel file to generate a "
        "live intervention report. Grades not yet entered are automatically excluded."
    )

    # Download blank template
    with open("/path/to/IBT_Student_Grade_Tracker.xlsx", "rb") as f:
        st.download_button(
            "📥 Download Blank Grade Tracker Template",
            f.read(),
            file_name="IBT_Student_Grade_Tracker.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Fill in student info and grades, then re-upload here."
        )

    uploaded = st.file_uploader(
        "Upload completed Grade Tracker (.xlsx)",
        type=["xlsx"],
        key="grade_tracker_upload",
        help="Upload the IBT Grade Tracker Excel file after entering grades."
    )

    if uploaded:
        with st.spinner("Ingesting grade data and computing averages…"):
            try:
                tracker = GradeTracker(uploaded)
                render_grade_report(tracker)
            except Exception as e:
                st.error(f"Could not parse the grade file: {e}\n\nMake sure you're using the IBT Grade Tracker template.")
